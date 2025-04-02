import warnings
from pathlib import Path
from openpyxl import load_workbook
from playwright.sync_api import Page
from controller.file import File
from helper.get_data_from_excel import coordinate_list_generator
from user_office.constants import MAIN_TAB

warnings.simplefilter(action='ignore', category=UserWarning)


class ExcelFillTemplate:

    def __init__(self, file):
        filename = load_workbook(filename=file)
        self.wb = filename
        self.ws = self.wb.active

    def collect_cells_in_dictionary(self):
        """Чтение 13 строки в шаблоне"""
        for row in self.ws.iter_rows(min_row=13, min_col=2, max_col=13, max_row=13, values_only=True):
            r = [list(row[i:i + 3]) for i in range(0, len(row), 3)]
            return r

    def collect_cells_in_excel(self, min_r=1, max_r=6):
        sheet = self.ws
        names = []
        for row in sheet.iter_rows(min_row=min_r, min_col=3, max_col=3, max_row=max_r, values_only=True):
            title = [list(row[i:i + 3]) for i in range(0, len(row), 3)]
            names.append(title)
        return names

    def fill_client_name(self, row, col, client_name):
        """Имя Клиента"""
        e = self.ws.cell(row=row, column=col)
        e.value = client_name

    def fill_campaign_name(self, row, col, campaign_name):
        """Имя кампании"""
        e = self.ws.cell(row=row, column=col)
        e.value = campaign_name

    def fill_channel(self, row, col, channel):
        """Канал Display"""
        e = self.ws.cell(row=row, column=col)
        e.value = channel

    def fill_seller(self, row, col, seller):
        """Селлер VK"""
        e = self.ws.cell(row=row, column=col)
        e.value = seller

    def fill_area(self, row, col, area):
        """Площадка MyTarget"""
        e = self.ws.cell(row=row, column=col)
        e.value = area

    def fill_targeting(self, row, col, text_targeting):
        """Колонока - таргетинга"""
        # row 13 col 5
        e = self.ws.cell(row=row, column=col)
        e.value = text_targeting

    def fill_geo_targeting(self, row, col, geo):
        """Колонока - геотаргетинга"""
        f = self.ws.cell(row=row, column=col)
        f.value = geo

    def fill_platform(self, row, col, platform):
        """Колонка платформы"""
        f = self.ws.cell(row=row, column=col)
        f.value = platform

    def fill_type_creative(self, row, col, type_creative):
        """Тип Креатива"""
        f = self.ws.cell(row=row, column=col)
        f.value = type_creative

    def fill_position_on_site(self, rows, col, position_site):
        """Колонка  - позиция на сайте"""
        f = self.ws.cell(row=rows, column=col)
        f.value = position_site

    def fill_format(self, row, col, format):
        """Формат 1000x100"""
        f = self.ws.cell(row=row, column=col)
        f.value = format

    def fill_period_placement(self):
        """Период размещения"""
        pass

    def fill_type_of_purchase(self, row, col, type_purchase):
        """Тип закупки"""
        f = self.ws.cell(row=row, column=col)
        f.value = type_purchase

    def fill_number_of_units_per_period(self, row, col, units_per_period):
        """Количество единиц за период"""
        f = self.ws.cell(row=row, column=col)
        f.value = units_per_period

    def fill_price_of_units_per_period(self, row, col, price_of_units_per_period):
        """Цена (за единицу покупки), руб."""
        f = self.ws.cell(row=row, column=col)
        f.value = price_of_units_per_period

    def fill_frequency(self, row, col, frequency):
        """Цена за единицу покупки"""
        f = self.ws.cell(row=row, column=col)
        f.value = frequency

    def fill_platform_price(self, row, col, platform_price):
        """Платформа"""
        f = self.ws.cell(row=row, column=col)
        f.value = platform_price

    def fill_geo_price(self, row, col, geo_price):
        """Цена - геотаргетинг"""
        f = self.ws.cell(row=row, column=col)
        f.value = geo_price

    def fill_age_sex_price(self, row, col, age_sex):
        """Цена - возраст и пол"""
        f = self.ws.cell(row=row, column=col)
        f.value = age_sex

    def fill_season_price(self, row, col, season):
        """Цена - сезонность"""
        f = self.ws.cell(row=row, column=col)
        f.value = season

    def fill_other_price(self, row, col, other_price):
        """Цена"""
        f = self.ws.cell(row=row, column=col)
        f.value = other_price

    def fill_currency_transfer(self, row, col, currency_transfer):
        """Цена процент скидки"""
        f = self.ws.cell(row=row, column=col)
        f.value = currency_transfer

    def fill_sum(self, row, col, sum):
        """Итого"""
        f = self.ws.cell(row=row, column=col)
        f.value = sum

    def fill_media_discount(self, row, col, media_discount):
        """Скидка медийная, %"""
        f = self.ws.cell(row=row, column=col)
        f.value = media_discount

    def fill_price_placement_on_ruble(self, row, col, price_placement_on_ruble):
        """Стоимость размещения, руб."""
        f = self.ws.cell(row=row, column=col)
        f.value = price_placement_on_ruble

    def fill_impressions(self, row, col, impressions):
        """Показы"""
        f = self.ws.cell(row=row, column=col)
        f.value = impressions

    def fill_coverage(self, row, col, coverage):
        """Охват"""
        f = self.ws.cell(row=row, column=col)
        f.value = coverage

    def fill_clicking_links(self, row, col, clicking_links):
        """Кликовая ссылка размещения"""
        f = self.ws.cell(row=row, column=col)
        f.value = clicking_links

    def save_template(self, file_path):
        """Сохранение заполненного шаблона, сохранение будет в виде нового файла"""
        self.wb.save(file_path)
        # self.wb.close()


class UploadExcel:
    def __init__(self, page: Page):
        self.page = page
        self.file_input = File(
            page,
            locator="data-testid='digital-media-plan-import-xls'",
            name='Загрузить медиплан',
        )

    def collect_cell(self):
        file = self.file_check_in_folder(
            folder='files',
            file_format='Темплейт_МП_ed2.4.xlsx',
            number=4,
        )
        excel = ExcelFillTemplate(file=file)
        cell = excel.collect_cells_in_dictionary()
        return cell

    def collect_cell_title(self):
        file = self.file_check_in_folder(
            folder='files',
            file_format='Темплейт_МП_ed2.4.xlsx',
            number=4,
        )
        excel = ExcelFillTemplate(file=file)
        cell_title = excel.collect_cells_in_excel()
        return cell_title

    @staticmethod
    def file_check_in_folder(
            folder: str, file_format: str, number: int
    ) -> Path:
        """Метод проверяющий наличие файлов в папке по формату
        number: принимает число, относительного пути для перехода по директории "вниз"
                например чтобы спустится на уровень ниже от папки automation/files в папку automation
                нужно передать в аргумнет цифру 1
        file_format: формат файла который нужен проверить в папке - например .xlsx
                file_format='.xlsx'
        folder: папка где нужно проверить искомый файл
                folder = './files'
        file_xlsx: путь до файла или абсолютно укзанный файл
        """

        current_path = Path.cwd().parents[number] / folder
        for file in current_path.iterdir():
            if file.suffix == file_format and file.is_file():
                assert file.is_file() == True, 'Файл отсутствует'
                assert (
                        file.suffix == file_format
                ), 'Расширение файла не соответсвует'
            return file

    def fill_excel_files(self, campaign_name, client_name):
        file = self.file_check_in_folder(
            folder='files',
            file_format='Темплейт_МП_ed2.4.xlsx',
            number=4,
        )
        excel = ExcelFillTemplate(file)
        excel.fill_client_name(row=1, col=3, client_name=client_name)
        excel.fill_campaign_name(row=2, col=3, campaign_name=campaign_name)
        excel.save_template(file_path=file)

    def upload_media_plan(self):
        file = self.file_check_in_folder(
            folder='files',
            file_format='Темплейт_МП_ed2.4.xlsx',
            number=4,
        )
        self.page.get_by_test_id('campaign-mediaplans_media-plan-import-xls').click()
        self.page.locator('.uppy-DragDrop-input').set_input_files(files=file)
        self.page.wait_for_load_state(timeout=3)

    def check_upload_excel_template(self):
        import_file = self.page.get_by_test_id('mediaplan-import-modal_file').inner_text()
        assert import_file == 'Темплейт_МП_ed2.4.xlsx\n\n0.01 MB', "Файл не загружен"
        self.page.click('text="Загрузить"')
